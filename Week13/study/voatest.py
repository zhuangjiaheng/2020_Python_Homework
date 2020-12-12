#-*- coding=utf-8 -*-
#@Time:  
#@Author: zjh
#@File: voatest.py
#@Software: PyCharm

import re
import json, requests, sys
from bs4 import BeautifulSoup
import openpyxl

file1 = r"./result/Essay.xlsx"

# 建立excel保存数据
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'English'
sheet.cell(row=1, column=1).value = 'Category'
sheet.cell(row=1, column=2).value = 'Essay'
sheet.cell(row=1, column=3).value = 'Link'
sheet.cell(row=1, column=4).value = 'Content'
sheet.cell(row=1, column=5).value = 'mp3_link'

# 打开网页，判断为静态网页
url = 'https://www.51voa.com/'
# params={'pageNo': '{}'.format(k),}
response = requests.get(url)
print(response.status_code)

# 爬取数据 并保存数据
pageSource = response.text  # 获取Elements中渲染完成的网页源代码
soup = BeautifulSoup(pageSource, 'html.parser')  # 使用bs解析网页
Essays_message = soup.find('div', id="list").find('ul').find_all('li')  # 使用bs提取元素
print(len(Essays_message))

i = 2
##导出文章列表
for Essay_message in Essays_message:
    Essay_Category = Essay_message.find_all('a')[0].text
    Essay_Essay = Essay_message.find_all('a')[-1].text
    Essay_Link = Essay_message.find_all('a')[-1]['href']

    # 保存数据
    sheet.cell(row=i, column=1).value = Essay_Category
    sheet.cell(row=i, column=2).value = Essay_Essay
    sheet.cell(row=i, column=3).value = 'https://www.51voa.com' + Essay_Link
    i = i + 1

    wb.save(file1)


def download_music(music_name, music_url):
    # """下载音乐"""
    response = requests.get(music_url)
    content = response.content
    save_file(r'D:\Python_E\others\trial\VOA\\' + music_name + '.mp3', content)


def save_file(filename, content):
    # """保存音乐"""
    with open(file=filename, mode="wb") as f:
        f.write(content)


##导出文章内容
rows = sheet.max_row
print(rows)

## 下面程式会报错，容易发现问题

for j in range(2,4+1):
    url = sheet.cell(row=j,column=3).value
    response = requests.get(url)
    # 爬取数据 并保存数据
    pageSource = response.text # 获取Elements中渲染完成的网页源代码
    soup = BeautifulSoup(pageSource,'html.parser')  # 使用bs解析网页
    Essay_contents = soup.find('div',id="Right_Content").find('div',class_="Content").find_all('p') # 使用bs提取元素
    Essay_content_list=[]
    for Essay_content in Essay_contents:
        Essay_content_p=Essay_content.text
        Essay_content_list.append(Essay_content_p)
        Essay_content="\n".join(Essay_content_list)
        sheet.cell(row=j,column=4).value=Essay_content

    #提取MP3链接
    Essay_MP3s = soup.find('a',id="mp3")['href'] # 使用bs提取元素
    sheet.cell(row=j,column=5).value=Essay_MP3s
    wb.save(file1)

    #更改歌名
    music_name=(sheet.cell(row=j,column=2).value).strip().replace('.', '').replace('?', '').replace('/', '').replace(' ', '').replace('(面议)','')
    music_url=sheet.cell(row=j,column=5).value
    download_music(music_name, music_url)

    # except Exception:
    #     pass

'''
## 下面程式不会报错，不容易发现问题
# for j in range(2,rows+1):
for j in range(2, 3 + 1):
    try:
        url = sheet.cell(row=j, column=3).value
        response = requests.get(url)
        # 爬取数据 并保存数据
        pageSource = response.text  # 获取Elements中渲染完成的网页源代码
        soup = BeautifulSoup(pageSource, 'html.parser')  # 使用bs解析网页
        Essay_contents = soup.find('div', id="Right_Content").find('div', class_="Content").find_all('p')  # 使用bs提取元素
        Essay_content_list = []
        for Essay_content in Essay_contents:
            Essay_content_p = Essay_content.text
            Essay_content_list.append(Essay_content_p)
            Essay_content = "\n".join(Essay_content_list)
            sheet.cell(row=j, column=4).value = Essay_content

        # 提取MP3链接
        Essay_MP3s = soup.find('a', id="mp3")['href']  # 使用bs提取元素
        sheet.cell(row=j, column=5).value = Essay_MP3s
        wb.save(file1)

        # 更改歌名
        music_name = (sheet.cell(row=j, column=2).value).strip().replace('.', '').replace('?', '').replace('/',
                                                                                                           '').replace(
            ' ', '_')
        music_url = sheet.cell(row=j, column=5).value
        download_music(music_name, music_url)

    except Exception:
        print(Exception)
        pass
'''
print('数据获取完毕，OK')
wb.close()